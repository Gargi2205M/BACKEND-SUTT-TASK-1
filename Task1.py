import json
import openpyxl
from collections import OrderedDict
from itertools import islice


wb = openpyxl.load_workbook('Names, Ids .xlsx')
sheet = wb.active
cell = sheet.cell(row = 1, column = 1)


sheet_name = wb.sheetnames
sheet = wb[sheet_name[0]]
branch = {'AA': 'ECE', 'AB': 'Manu', 'A1': 'Chemical', 'A2': 'Civil', 'A3': 'EEE', 'A4': 'Mech', 'A5': 'Pharma', 'A7': 'CSE', 'A8': 'ENI', 'B1': 'MSc BIO', 'B2': 'MSc Chem', 'B3': 'MSc Eco', 'B4': 'MSc Mathematics', 'B5': 'MSc Physics'}


        
students_list = []
for row in islice(sheet.values, 1, sheet.max_row):
    if row != (None, None):
        students = OrderedDict()
        students['name'] = row[1]
        students['bits-id'] = row[0]
        students['bits-email'] = 'F' + '2022' + (row[0])[8:12] + '@pilani.bits-pilani.ac.in'
        students['branch'] = branch[(row[0])[4:6]]
        students_list.append(students)
    


with open('data.json', 'w') as f:
    j = json.dumps(students_list)
    f.write(j)



    
    
    
    

    

    
    


